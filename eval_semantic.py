import argparse
from pathlib import Path
import numpy as np
from PIL import Image
import openpyxl

def read_and_resize_labels(path, size):
    image = Image.open(path)
    return np.array(image.resize(size, Image.NEAREST))

def get_non_robust_classes(confusion_matrix, robustness_thres):
    axis_0 = np.sum(confusion_matrix, axis=0)
    axis_1 = np.sum(confusion_matrix, axis=1)
    total_labels = axis_0.sum()
    non_robust_0 = axis_0 / total_labels < robustness_thres
    non_robust_1 = axis_1 / total_labels < robustness_thres
    return np.where(np.logical_and(non_robust_0, non_robust_1))[0].tolist()

def calculate_miou(confusion_matrix, ignore_class=None, robust=0.005):
    MIoU = np.divide(np.diag(confusion_matrix), (np.sum(confusion_matrix, axis=1) + np.sum(confusion_matrix, axis=0) - np.diag(confusion_matrix)))
    if ignore_class is not None:
        ignore_class += get_non_robust_classes(confusion_matrix, robust)
        for i in ignore_class:
            MIoU[i] = float('nan')
    MIoU = np.nanmean(MIoU)
    return MIoU

class ConfusionMatrix:
    def __init__(self, num_classes, ignore_class=None, robust=0.005):
        np.seterr(divide='ignore', invalid='ignore')
        self.num_class = num_classes
        self.ignore_class = ignore_class
        self.confusion_matrix = np.zeros((self.num_class,) * 2)
        self.robust = robust

    def _generate_matrix(self, gt_image, pre_image):
        mask = (gt_image >= 0) & (gt_image < self.num_class)
        label = self.num_class * gt_image[mask].astype('int') + pre_image[mask]
        count = np.bincount(label, minlength=self.num_class**2)
        confusion_matrix = count.reshape(self.num_class, self.num_class)
        return confusion_matrix

    def add_batch(self, gt_image, pre_image, return_miou=False):
        assert gt_image.shape == pre_image.shape
        confusion_matrix = self._generate_matrix(gt_image, pre_image)
        self.confusion_matrix += confusion_matrix
        if return_miou:
            return calculate_miou(confusion_matrix, self.ignore_class, self.robust)

    def get_miou(self):
        return calculate_miou(self.confusion_matrix, self.ignore_class, self.robust)

    def calculate_metrics(self):
        total_pixels = np.sum(self.confusion_matrix)
        tp = np.diag(self.confusion_matrix)
        fp = np.sum(self.confusion_matrix, axis=0) - tp
        fn = np.sum(self.confusion_matrix, axis=1) - tp

        iou = tp / (tp + fp + fn)
        class_avg_accuracy = tp / (tp + fn)
        total_accuracy = np.sum(tp) / total_pixels
        fw_iou = (np.sum(tp + fn) / total_pixels) * np.sum(iou)

        return iou, class_avg_accuracy, total_accuracy, fw_iou

    def reset(self):
        self.confusion_matrix = np.zeros((self.num_class,) * 2)



def calculate_iou(y_true, y_pred, classes):
    iou = []
    for class_id in classes:
        intersection = np.sum((y_true == class_id) & (y_pred == class_id))
        union = np.sum((y_true == class_id) | (y_pred == class_id))
        iou.append(intersection / union if union > 0 else 0)
    return iou

def calculate_miou(y_true, y_pred, classes):
    class_iou = calculate_iou(y_true, y_pred, classes)
    return np.mean(class_iou)

def calculate_fwiou(y_true, y_pred, classes):
    class_iou = calculate_iou(y_true, y_pred, classes)
    class_frequency = np.sum(y_true == np.arange(classes), axis=(0, 1))
    class_weight = class_frequency / np.sum(class_frequency)
    fw_iou = np.sum(class_iou * class_weight)
    return fw_iou

def calculate_class_average_accuracy(y_true, y_pred, classes):
    class_accuracy = []
    for class_id in classes:
        tp = np.sum((y_true == class_id) & (y_pred == class_id))
        fn = np.sum((y_true == class_id) & (y_pred != class_id))
        class_accuracy.append(tp / (tp + fn + 1e-10))  # Add a small epsilon to avoid division by zero
    return np.mean(class_accuracy)

def calculate_total_accuracy(y_true, y_pred):
    tp = np.sum(y_true == y_pred)
    total_pixels = y_true.size
    return tp / total_pixels


class Map():
    def __init__(self):

        xlsx_file_path = '/mnt/user/datasets/replica/semantic2color.xlsx'

        self.semantic2color = {}
        workbook = openpyxl.load_workbook(xlsx_file_path)
        sheet = workbook.active
        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, values_only=True):
            self.semantic2color.update({str(col[1:])[1:4]: col[0]})

        self.v_map_function = np.vectorize(self.map_function)


    def map_function(self, value):
            label = self.semantic2color.get(str(value)[1:4], 0)
            return label


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='metrics')
    args = parser.parse_args()

    rooms = [
        'room_0',
        'room_1',
        'room_2',
        'office_0',
        'office_1',
        'office_2',
        'office_3',
        'office_4',
    ]

    import ipdb
    ipdb.set_trace()

    color2label = Map()

    for room in rooms:
        miou = []
        fwiou = []
        class_avg_accuracy = []
        total_accuracy = []

        for idx in range(0, 2000, 10):
                
            pred = f'/mnt/user/so_nice_slam_1112/checkpoints/replica_sdf10/{room}/rendered/semantic_vis_{idx:05d}.png'
            gt = f'/mnt/user/so_nice_slam_1112/checkpoints/replica_sdf10/{room}/rendered/semantic_gt_vis_{idx:05d}.png'
            pred = read_and_resize_labels(pred, (1200, 680))
            gt = read_and_resize_labels(gt, (1200, 680))

            y_pred = color2label.v_map_function(pred)
            y_true = color2label.v_map_function(gt)

            classes = np.unique(y_true)
            miou.append(calculate_miou(y_true, y_pred, classes))
            fwiou.append(calculate_fwiou(y_true, y_pred, classes))
            class_avg_accuracy.append(calculate_class_average_accuracy(y_true, y_pred, classes))
            total_accuracy.append(calculate_total_accuracy(y_true, y_pred))

        print(f"mIoU: {miou}")
        print(f"fwIoU: {fwiou}")
        print(f"Class Average Accuracy: {class_avg_accuracy}")
        print(f"Total Accuracy: {total_accuracy}")
