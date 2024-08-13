import time, os
import math
import argparse
import numpy
import copy
import torch
import sys
import numpy as np
import yaml
import cv2
import openpyxl

from utils.common import load_config
from slams.meshing import Vis
from models.encoder import ResNet

import datas as data
from models.checkpoint import Checkpoint
from models.decoder import Decoder
from slams.meshing import Mesher


class DNS_SLAM():
    def __init__(self, cfg, device):

        self.cfg = cfg
        self.device = device[1]

        self.input_dir = cfg['dataset_dir'] + '/' + cfg['dataset'] + '/' + cfg['scene']
        self.out_dir = cfg['out_dir'] + '/' + cfg['dataset'] + '/' + cfg['scene']   
        if not os.path.exists(self.out_dir):
            os.makedirs(self.out_dir)

        self.low_gpu_mem = cfg['low_gpu_mem']
        self.verbose = cfg['verbose']
        self.scale = cfg['scale']

        self.H, self.W= cfg['cam']['H'], cfg['cam']['W']
        self.fx = cfg['cam']['fx']
        self.fy = cfg['cam']['fy']
        self.cx = cfg['cam']['cx']
        self.cy = cfg['cam']['cy']
        self.update_cam()

        self.K = torch.tensor([[self.fx,  0,     self.cx],
                               [0,     self.fy,  self.cy],
                               [0,        0,        1   ]]).to(self.device)

        self.dataset = data.get_dataset(cfg, self.input_dir, 1, self.device)
        self.n_class = self.dataset.n_class
        self.n_img = self.dataset.n_img
        self.class2label_dict = self.dataset.class2label_dict

        self.n_surface_ray = cfg['training']['n_surface_ray']
        self.n_samples_ray = cfg['training']['n_samples_ray']
        self.pts_dim = cfg['model']['pts_dim']
        self.hidden_dim = cfg['model']['hidden_dim']
        self.pixel_dim = cfg['model']['pixel_dim']

        # Model initialization
        self.load_bound()
        self.decoder = Decoder(cfg['model'], self.bound.to(self.device), self.n_class).to(self.device)
        
        self.lr = cfg['training']['lr']

        self.checkpoint = Checkpoint(self.out_dir, device=self.device,
                                     decoder=self.decoder)
        
        if cfg['dataset'] == 'replica':
            self.v_map_function = np.vectorize(self.map_function)
            xlsx_file_path = cfg['dataset_dir'] + '/' + cfg['dataset'] + '/semantic2color.xlsx'

            self.semantic2color = {}
            workbook = openpyxl.load_workbook(xlsx_file_path)
            sheet = workbook.active
            for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, values_only=True):
                self.semantic2color.update({col[0]: col[1:]})
        else:
            self.v_map_function = None

        self.visualizer = Mesher(self.cfg, self, device=self.device)

    def map_function(self, value):
        replica_class = self.class2label_dict.get(value, value)
        color = self.semantic2color.get(replica_class, [0, 0, 0])
        return color  # .get(value, default)

    def load_bound(self):
        # scale the bound if there is a global scaling factor
        self.bound = torch.from_numpy(
            np.array(self.cfg['back_end']['bound'])*self.scale).to(self.device)
        bound_divisible = self.cfg['bound_divisible']
        # enlarge the bound a bit to allow it divisible by bound_divisible
        self.bound[:, 1] = (((self.bound[:, 1]-self.bound[:, 0]) /
                            bound_divisible).int()+1)*bound_divisible+self.bound[:, 0]

    
    def update_cam(self):
        """
        Update the camera intrinsics according to pre-processing config, 
        such as resize or edge crop.
        """
        # resize the input images to crop_size (variable name used in lietorch)
        if 'crop_size' in self.cfg['cam']:
            crop_size = self.cfg['cam']['crop_size']
            sx = crop_size[1] / self.W
            sy = crop_size[0] / self.H
            self.fx = sx*self.fx
            self.fy = sy*self.fy
            self.cx = sx*self.cx
            self.cy = sy*self.cy
            self.W = crop_size[1]
            self.H = crop_size[0]

        # croping will change H, W, cx, cy, so need to change here
        if self.cfg['cam']['crop_edge'] > 0:
            self.H -= self.cfg['cam']['crop_edge']*2
            self.W -= self.cfg['cam']['crop_edge']*2
            self.cx -= self.cfg['cam']['crop_edge']
            self.cy -= self.cfg['cam']['crop_edge']


if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Generating mesh.'
    )
    parser.add_argument('config', type=str, help='Path to config file.')
    parser.add_argument('--input_folder', type=str,
                        help='input folder, this have higher priority, can overwrite the one in config file')
    parser.add_argument('--output', type=str,
                        help='output folder, this have higher priority, can overwrite the one in config file')
    nice_parser = parser.add_mutually_exclusive_group(required=False)
    parser.set_defaults(nice=True)
    args = parser.parse_args()

    cfg = load_config(args.config, 'configs/slam.yaml')

    device = [torch.device('cuda:0'), torch.device('cuda:0')]
    slam = DNS_SLAM(cfg, device)
    output = slam.out_dir

    ckpt = slam.checkpoint.load('model.pt')
    N = ckpt['idx']
    encoder = ResNet().to(device[1])
    decoder = slam.decoder
    slam.fine_decoders = ckpt['fine_decoders']
    fine_decoders = ckpt['fine_decoders']
    keyframe_list = ckpt['keyframe_list']
    keyframe_dict = ckpt['keyframe_dict']
    estimate_c2w_list = ckpt['estimate_c2w_list']

    # mesh
    mesh_out_file = output
    slam.visualizer.get_mesh(mesh_out_file,
                            encoder, 
                            decoder, 
                            fine_decoders,
                            keyframe_dict, 
                            estimate_c2w_list, 
                            N, 
                            device[1], 
                            color=True,
                            label=True,
                            element=False,
                            clean_mesh=True,
                            get_mask_use_all_frames=False)


