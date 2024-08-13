import time, os
import numpy as np
import openpyxl
import torch
import torch.multiprocessing as mp
torch.multiprocessing.set_sharing_strategy('file_system')

import datas as data
from models.checkpoint import Checkpoint
from models.decoder import Decoder
from slams.tracking import Tracker
from slams.mapping import Mapper
from slams.meshing import Mesher


class DNS_SLAM():
    def __init__(self, cfg, device):

        self.cfg = cfg
        self.device = device
        self.scene = cfg['scene']   
        self.input_dir = cfg['dataset_dir'] + '/' + cfg['dataset'] + '/' + cfg['scene']
        self.out_dir = cfg['out_dir'] + '/' + cfg['dataset'] + '/' + cfg['scene']    
        if not os.path.exists(self.out_dir):
            os.makedirs(self.out_dir)

        self.low_gpu_mem = cfg['low_gpu_mem']
        self.verbose = cfg['verbose']
        self.scale = cfg['scale']

        self.H, self.W= cfg['cam']['H'], cfg['cam']['W']
        self.fx = cfg['cam']['fx']
        self.fy = cfg['cam']['fy']
        self.cx = cfg['cam']['cx']
        self.cy = cfg['cam']['cy']
        self.update_cam()

        self.front_dataset = data.get_dataset(cfg, self.input_dir, 1, self.device[0])
        self.front_loader = torch.utils.data.DataLoader(
            self.front_dataset, batch_size=1, num_workers=1, shuffle=False)
        self.back_dataset = data.get_dataset(cfg, self.input_dir, 1, self.device[1])
        self.n_class = self.back_dataset.n_class

        # Model initialization
        self.load_bound()
        self.decoder = Decoder(cfg['model'], self.bound.to(device[1]), self.n_class).to(device[1])
        self.decoder.share_memory()
        
        self.lr = cfg['training']['lr']

        self.checkpoint = Checkpoint(self.out_dir, device=device[1],
                                     decoder=self.decoder)
        
        self.front_idx = torch.ones((1)).int()
        self.front_idx.share_memory_()
        self.back_idx = torch.zeros((1)).int()
        self.back_idx.share_memory_()
        self.optimize_every_frame = torch.ones((1)).int()

        self.n_img = len(self.front_loader)
        self.estimate_c2w_list = torch.zeros((self.n_img, 4, 4))
        self.estimate_c2w_list.share_memory_()
        self.gt_c2w_list = torch.zeros((self.n_img, 4, 4))
        self.gt_c2w_list.share_memory_()

        self.first_frame_optimized = torch.zeros((1)).int()
        self.first_frame_optimized.share_memory_()

        if cfg['dataset'] == 'replica':
            self.v_map_function = np.vectorize(self.map_function)
            xlsx_file_path = cfg['dataset_dir'] + '/' + cfg['dataset'] + '/semantic2color.xlsx'

            self.semantic2color = {}
            workbook = openpyxl.load_workbook(xlsx_file_path)
            sheet = workbook.active
            for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, values_only=True):
                self.semantic2color.update({col[0]: col[1:]})
        else:
            self.v_map_function = None

        self.mesher = Mesher(self.cfg, self, device=self.device[1])
        self.front = Tracker(self.cfg, self, self.device[0])
        self.back = Mapper(self.cfg, self, self.device[1])

    def map_function(self, value):
        replica_class = self.class2label_dict.get(value, value)
        color = self.semantic2color.get(replica_class, [0, 0, 0])
        return color

    def load_bound(self):
        # scale the bound if there is a global scaling factor
        self.bound = torch.from_numpy(
            np.array(self.cfg['back_end']['bound'])*self.scale)
        bound_divisible = self.cfg['bound_divisible']
        # enlarge the bound a bit to allow it divisible by bound_divisible
        self.bound[:, 1] = (((self.bound[:, 1]-self.bound[:, 0]) /
                            bound_divisible).int()+1)*bound_divisible+self.bound[:, 0]

    
    def update_cam(self):
        """
        Update the camera intrinsics according to pre-processing config, 
        such as resize or edge crop.
        """
        # resize the input images to crop_size (variable name used in lietorch)
        if 'crop_size' in self.cfg['cam']:
            crop_size = self.cfg['cam']['crop_size']
            sx = crop_size[1] / self.W
            sy = crop_size[0] / self.H
            self.fx = sx*self.fx
            self.fy = sy*self.fy
            self.cx = sx*self.cx
            self.cy = sy*self.cy
            self.W = crop_size[1]
            self.H = crop_size[0]

        # croping will change H, W, cx, cy, so need to change here
        if self.cfg['cam']['crop_edge'] > 0:
            self.H -= self.cfg['cam']['crop_edge']*2
            self.W -= self.cfg['cam']['crop_edge']*2
            self.cx -= self.cfg['cam']['crop_edge']
            self.cy -= self.cfg['cam']['crop_edge']

    
    def front_end(self, rank):
        '''
        Tracking Thread.
        '''
        while (1):
            if self.first_frame_optimized[0] == 1:
                break
            time.sleep(1)
        
        self.front.run()


    def back_end(self, rank):
        """
        Mapping Thread.
        """
        self.back.run()


    def run(self):
        # try:
        #     load_dict = self.checkpoint.load('model.pt')
        # except FileNotFoundError:
        #     load_dict = dict()
        #     print('Fail to load, initializing...')

        processes = []
        for rank in range(2):
            if rank == 0:
                p = mp.Process(target=self.back_end, args=(rank, ))
            else:
                p = mp.Process(target=self.front_end, args=(rank, ))
            
            p.start()
            # p.run()
            processes.append(p)
        for p in processes:
            p.join()


if __name__ == '__main__':
    pass
