import time, os
import math
import argparse
import numpy
import copy
import torch
import sys
import numpy as np
import yaml
import cv2
import openpyxl
from pytorch_msssim import ms_ssim
from torchmetrics.image.lpip import LearnedPerceptualImagePatchSimilarity

from utils.common import load_config
from slams.meshing import Vis
from models.encoder import ResNet

import datas as data
from models.checkpoint import Checkpoint
from models.decoder import Decoder
from utils.common import get_all_rays, sample_along_rays, feature_matching, fig_plot, raw2nerf_color


class DNS_SLAM():
    def __init__(self, cfg, device):

        self.cfg = cfg
        self.device = device[1]

        self.dataset_ = cfg['dataset']
        self.scene = cfg['scene']
        self.input_dir = cfg['dataset_dir'] + '/' + self.dataset_ + '/' + self.scene
        self.out_dir = cfg['out_dir'] + '/' + self.dataset + '/' + self.scene       
        if not os.path.exists(self.out_dir):
            os.makedirs(self.out_dir)

        self.low_gpu_mem = cfg['low_gpu_mem']
        self.verbose = cfg['verbose']
        self.scale = cfg['scale']

        self.H, self.W= cfg['cam']['H'], cfg['cam']['W']
        self.fx = cfg['cam']['fx']
        self.fy = cfg['cam']['fy']
        self.cx = cfg['cam']['cx']
        self.cy = cfg['cam']['cy']
        self.update_cam()

        self.K = torch.tensor([[self.fx,  0,     self.cx],
                               [0,     self.fy,  self.cy],
                               [0,        0,        1   ]]).to(self.device)

        self.dataset = data.get_dataset(cfg, self.input_dir, 1, self.device)
        self.n_class = self.dataset.n_class
        self.n_img = self.dataset.n_img
        self.class2label_dict = self.dataset.class2label_dict

        self.n_surface_ray = cfg['training']['n_surface_ray']
        self.n_samples_ray = cfg['training']['n_samples_ray']
        self.pts_dim = cfg['model']['pts_dim']
        self.hidden_dim = cfg['model']['hidden_dim']
        self.pixel_dim = cfg['model']['pixel_dim']

        # Model initialization
        self.load_bound()
        self.decoder = Decoder(cfg['model'], self.bound.to(self.device), self.n_class).to(self.device)
        
        self.lr = cfg['training']['lr']

        self.checkpoint = Checkpoint(self.out_dir, device=self.device,
                                     decoder=self.decoder)
        
        self.v_map_function = np.vectorize(self.map_function)

        self.n_point_batch = 8000

        self.psnr_sum = 0
        self.ssim_sum = 0
        self.lpips_sum = 0

        self.cal_lpips = LearnedPerceptualImagePatchSimilarity(
                    net_type='alex', normalize=True).to(self.device)

        xlsx_file_path = cfg['dataset_dir'] + '/' + self.dataset_ + '/semantic2color.xlsx'

        self.semantic2color = {}
        workbook = openpyxl.load_workbook(xlsx_file_path)
        sheet = workbook.active
        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, values_only=True):
            self.semantic2color.update({col[0]: col[1:]})


    def map_function(self, value):
        replica_class = self.class2label_dict.get(value, value)
        color = self.semantic2color.get(replica_class, [0, 0, 0])
        return color  # .get(value, default)

    def load_bound(self):
        # scale the bound if there is a global scaling factor
        self.bound = torch.from_numpy(
            np.array(self.cfg['back_end']['bound'])*self.scale).to(self.device)
        bound_divisible = self.cfg['bound_divisible']
        # enlarge the bound a bit to allow it divisible by bound_divisible
        self.bound[:, 1] = (((self.bound[:, 1]-self.bound[:, 0]) /
                            bound_divisible).int()+1)*bound_divisible+self.bound[:, 0]

    
    def update_cam(self):
        """
        Update the camera intrinsics according to pre-processing config, 
        such as resize or edge crop.
        """
        # resize the input images to crop_size (variable name used in lietorch)
        if 'crop_size' in self.cfg['cam']:
            crop_size = self.cfg['cam']['crop_size']
            sx = crop_size[1] / self.W
            sy = crop_size[0] / self.H
            self.fx = sx*self.fx
            self.fy = sy*self.fy
            self.cx = sx*self.cx
            self.cy = sy*self.cy
            self.W = crop_size[1]
            self.H = crop_size[0]

        # croping will change H, W, cx, cy, so need to change here
        if self.cfg['cam']['crop_edge'] > 0:
            self.H -= self.cfg['cam']['crop_edge']*2
            self.W -= self.cfg['cam']['crop_edge']*2
            self.cx -= self.cfg['cam']['crop_edge']
            self.cy -= self.cfg['cam']['crop_edge']

    def fine_fn(self, pes, classes=None, features=None):
        n_class = torch.unique(classes, sorted=True).cpu().int().numpy()
        latents = torch.zeros(pes.shape[0], self.hidden_dim+1, device=pes.device)
        for i in n_class:
            if i not in self.fine_decoders.keys():
                raise ValueError('Fine decoders does NOT have class', i)
            index = (classes[:] == i).cpu().numpy()
            if index.sum() > 1:
                pe = pes[index, :]
                feature = features[index, :]
                # latents[index, :] = self.fine_decoders[i](torch.cat((pe, feature), -1)).float()
                latents[index, :] = self.fine_decoders[i](pe, features=feature).float()
        return latents
    
    def renderer(self, samples):
        pts = samples['pts']  # [B*n_pts, n_sample, 3]
        n_pts, n_samples, _  = pts.shape
        pts = pts.flatten(0, 1)  # [B*N*n_samples, C]
        pts = (pts - self.bound[:, 0]) / (self.bound[:, 1] - self.bound[:, 0])

        rays_d = samples['rays_d']  # [B*n_pts, 3]
        z_vals = samples['z_vals']  # [B*n_pts, n_sample]
        gt_label = samples['gt_label']
        gt_label = gt_label.repeat(1, n_samples).flatten(0, 1)

        # reference feature
        pixel_pts = samples['features']
        pixel_pts = pixel_pts.flatten(0, 1)

        pe, grid_pts, fine_grid_pts = self.decoder.pe_fn(pts)

        coarse_latents = self.decoder.coarse_fn(pe, features=grid_pts)
        fine_latents = self.fine_fn(pe, classes=gt_label, features=grid_pts)
        # fine_latents = None

        color_pts, logits_pts = self.decoder.out_fn(pe, torch.cat((fine_latents[:, 1:], pixel_pts), -1))

        values_pts = torch.cat((color_pts, fine_latents[:, 0:1]), -1)

        values_pts = values_pts.reshape(n_pts, n_samples, -1)
        logits_pts = logits_pts.reshape(n_pts, n_samples, -1)

        pred_depth, pred_depth_var, pred_color, weights = raw2nerf_color(
            values_pts, z_vals, rays_d, device=self.device)
        pred_logits = torch.sum(weights[..., None] * logits_pts, -2)  # (N_rays, 3)

        return pred_color, pred_depth, pred_depth_var, pred_logits, fine_latents, coarse_latents

    def calculate_iou(self, y_true, y_pred, classes):
        iou = []
        for class_id in classes:
            intersection = np.sum((y_true == class_id) & (y_pred == class_id))
            union = np.sum((y_true == class_id) | (y_pred == class_id))
            iou.append(intersection / union if union > 0 else 0)
        return iou

    def calculate_miou(self, y_true, y_pred, classes):
        class_iou = self.calculate_iou(y_true, y_pred, classes)
        return np.mean(class_iou)

    def calculate_fwiou(self, y_true, y_pred, classes):
        class_iou = self.calculate_iou(y_true, y_pred, classes)
        class_frequency = []
        for class_id in classes:
            class_frequency.append(np.sum(y_true == class_id))
        class_weight = class_frequency / np.sum(class_frequency)
        fw_iou = np.sum(class_iou * class_weight)
        return fw_iou

    def calculate_class_average_accuracy(self, y_true, y_pred, classes):
        class_accuracy = []
        for class_id in classes:
            tp = np.sum((y_true == class_id) & (y_pred == class_id))
            fn = np.sum((y_true == class_id) & (y_pred != class_id))
            class_accuracy.append(tp / (tp + fn + 1e-10))  # Add a small epsilon to avoid division by zero
        return np.mean(class_accuracy)

    def calculate_total_accuracy(self, y_true, y_pred):
        tp = np.sum(y_true == y_pred)
        total_pixels = y_true.size
        return tp / total_pixels
    

    def novel_view_render(self, idx, cur_gt_color, cur_gt_depth, 
        cur_gt_label, cur_gt_c2w, cur_c2w, kf_idx, kf_list, kf_dict, encoder):
        with torch.no_grad():            
            H, W, _ = cur_gt_color.shape

            first = max(0, kf_idx - 1)
            second = min(len(kf_dict)-1, kf_idx + 1)
            refer_frame_idx = [first, second]

            refer_color_list = []
            refer_c2w_list = []
            for refer_id in refer_frame_idx:
                gt_color = kf_dict[refer_id]['gt_color'].to(self.device)
                est_c2w = kf_dict[refer_id]['est_c2w'].to(self.device)

                refer_color_list.append(gt_color.float())
                refer_c2w_list.append(est_c2w.float())
            
            refer_color_list = torch.stack(refer_color_list, 0).to(self.device)
            refer_c2w_list = torch.stack(refer_c2w_list, 0).to(self.device)

            # refer_frames = {'gt_color': torch.cat((refer_color_list, cur_gt_color.unsqueeze(0)), 0).unsqueeze(0).to(self.device),  # [B, N, H, W, 3]
            #             'est_c2w': torch.cat((refer_c2w_list, cur_c2w.unsqueeze(0)), 0).unsqueeze(0).to(self.device)}  # kf_idx: which kf is this reference frame from

            refer_frames = {'gt_color': cur_gt_color.unsqueeze(0).unsqueeze(0).to(self.device),  # [B, N, H, W, 3]
                        'est_c2w': cur_c2w.unsqueeze(0).unsqueeze(0).to(self.device)}  # kf_idx: which kf is this reference frame from

            rays_o, rays_d = get_all_rays(self.H, self.W, 
                                self.fx, self.fy, self.cx, self.cy, cur_c2w, self.device)
            rays_o = rays_o.flatten(0, 1)  # [N, 3]
            rays_d = rays_d.flatten(0, 1)
            cur_gt_label = cur_gt_label.flatten(0, 1)

            det_rays_o = rays_o.clone().detach().unsqueeze(-1)  # (N, 3, 1)
            det_rays_d = rays_d.clone().detach().unsqueeze(-1)  # (N, 3, 1)
            t = (self.bound.unsqueeze(0).to(self.device) -
                det_rays_o)/det_rays_d  # (N, 3, 2)
            far_bb, _ = torch.min(torch.max(t, dim=2)[0], dim=1)
            far_bb = far_bb.unsqueeze(-1)
            far_bb += 0.01
            
            z_vals = sample_along_rays(cur_gt_depth.flatten(0, 1), self.n_samples_ray, self.n_surface_ray,  
                                       far_bb, self.device)  # [n_pixels, n_samples]
            pts = rays_o[:, None, :] + rays_d[:, None, :] * z_vals[:, :, None]

            features = encoder(refer_frames['gt_color'][-1, ...].unsqueeze(0))  # [1, n_class, C]

            n_pts = pts.shape[0]
            n_split = math.ceil(n_pts / self.n_point_batch)        
            pred_color = []
            pred_depth = []  
            pred_logits = []
            refer_c2w = refer_frames['est_c2w'][0]
            w2c = torch.inverse(refer_c2w)
            for i in range(n_split):
                start = i * self.n_point_batch
                end = min((i + 1) * self.n_point_batch, n_pts) 

                pts_sp = pts[start:end, :, :]
                code = feature_matching(
                    self.H, self.W, self.K, pts_sp.flatten(0, 1), w2c, features[0], self.decoder.merge)
                code = code.reshape(pts_sp.shape[0], pts_sp.shape[1], -1)

                samples = {'rays_o': rays_o[start:end, :],  # [N, 3]
                        'rays_d': rays_d[start:end, :],  # [N, 3]
                        'gt_label': cur_gt_label[start:end],
                        'pts': pts_sp,
                        'z_vals': z_vals[start:end, :],
                        'features': code}  # [N, n_samples]
                
                color, depth, _, logits, _, _ = self.renderer(samples)  # [N, n_samples, C], [N, n_samples, num_slots]
                pred_color.append(color)
                pred_depth.append(depth)
                pred_logits.append(logits)

            pred_color = torch.cat(pred_color, dim = 0)
            pred_depth = torch.cat(pred_depth, dim = 0)
            pred_logits = torch.cat(pred_logits, dim = 0)
            pred_label = torch.argmax(pred_logits, dim=-1)

            pred_color = pred_color.reshape(H, W, -1)
            pred_depth = pred_depth.reshape(H, W)
            pred_label = pred_label.reshape(H, W)

            mse_loss = torch.nn.functional.mse_loss(
                cur_gt_color[cur_gt_depth > 0], pred_color[cur_gt_depth > 0])
            psnr_frame = -10. * torch.log10(mse_loss)
            ssim_value = ms_ssim(cur_gt_color.transpose(0, 2).unsqueeze(0).float(), pred_color.transpose(0, 2).unsqueeze(0).float(),
                                    data_range=1.0, size_average=True)
            lpips_value = self.cal_lpips(torch.clamp(cur_gt_color.unsqueeze(0).permute(0, 3, 1, 2).float(), 0.0, 1.0),
                                    torch.clamp(pred_color.unsqueeze(0).permute(0, 3, 1, 2).float(), 0.0, 1.0)).item()
            self.psnr_sum += psnr_frame
            self.ssim_sum += ssim_value
            self.lpips_sum += lpips_value

            print(f'{idx}th rendering result: psnr={psnr_frame}, ssim={ssim_value}, lpips={lpips_value}')

            if not os.path.exists(f'{self.out_dir}/rendered/'):
                os.makedirs(f'{self.out_dir}/rendered/')
            file = open(f'{self.out_dir}/rendered/rendering_eval.txt', 'a')
            file.write(f'Current frame: {idx} ' +
                        f'psnr: {psnr_frame:.4f} ' + 
                        f'ssim: {ssim_value:.4f} ' +
                        f'lpips: {lpips_value:.4f}\n')
            file.close()

            gt_color_np = cur_gt_color.cpu().numpy()
            gt_depth_np = cur_gt_depth.cpu().numpy()
            gt_label_np = cur_gt_label.reshape(H, W).cpu().numpy()
            color_np = pred_color.detach().cpu().numpy()
            depth_np = pred_depth.detach().cpu().numpy()
            label_np = pred_label.detach().cpu().numpy()

            classes = np.unique(gt_label_np)
            miou = self.calculate_miou(gt_label_np, label_np, classes)
            fwiou = self.calculate_fwiou(gt_label_np, label_np, classes)
            class_avg_accuracy = self.calculate_class_average_accuracy(gt_label_np, label_np, classes)
            total_accuracy = self.calculate_total_accuracy(gt_label_np, label_np)

            print(f'miou={miou}, fwiou={fwiou}, class_avg_acc={class_avg_accuracy}, total_acc={total_accuracy}')

            label_np = self.v_map_function(label_np)
            gt_label_np = self.v_map_function(gt_label_np)

            label_vis_np = np.zeros_like(gt_color_np)
            gt_label_vis_np = np.zeros_like(gt_color_np)

            for i in range(3):
                label_vis_np[..., i] = label_np[i]
                gt_label_vis_np[..., i] = gt_label_np[i]

            img = cv2.cvtColor(
                    color_np*255, cv2.COLOR_BGR2RGB)
            cv2.imwrite(f'{self.out_dir}/rendered/color_{idx:05d}.png', img)

            semantic = cv2.cvtColor(
                    label_vis_np, cv2.COLOR_BGR2RGB)
            cv2.imwrite(f'{self.out_dir}/rendered/semantic_vis_{idx:05d}.png', semantic)

            gt_semantic = cv2.cvtColor(
                    gt_label_vis_np, cv2.COLOR_BGR2RGB)
            cv2.imwrite(f'{self.out_dir}/rendered/semantic_gt_vis_{idx:05d}.png', gt_semantic)

        return miou, fwiou, class_avg_accuracy, total_accuracy



if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        description='Generating mesh.'
    )
    parser.add_argument('config', type=str, help='Path to config file.')
    parser.add_argument('--input_folder', type=str,
                        help='input folder, this have higher priority, can overwrite the one in config file')
    parser.add_argument('--output', type=str,
                        help='output folder, this have higher priority, can overwrite the one in config file')
    nice_parser = parser.add_mutually_exclusive_group(required=False)
    parser.set_defaults(nice=True)
    args = parser.parse_args()

    cfg = load_config(args.config, 'configs/dns_slam.yaml')

    device = [torch.device('cuda:0'), torch.device('cuda:0')]
    slam = DNS_SLAM(cfg, device)
    output = slam.out_dir

    ckpt = slam.checkpoint.load('model.pt')
    N = ckpt['idx']
    encoder = ResNet().to(device[1])
    decoder = slam.decoder
    slam.fine_decoders = ckpt['fine_decoders']
    fine_decoders = ckpt['fine_decoders']
    keyframe_list = ckpt['keyframe_list']
    keyframe_dict = ckpt['keyframe_dict']
    estimate_c2w_list = ckpt['estimate_c2w_list']


    # render
    count = 0
    miou_list = []
    fwiou_list = []
    class_avg_accuracy_list = []
    total_accuracy_list = []
    for i in range(N):
        # if i % cfg['back_end']['optimize_every_n_frames'] == 0: 
        if i % 10 == 0: 
            count += 1
            idx, gt_color, gt_depth, gt_label, gt_c2w = slam.dataset[i]
            est_c2w = estimate_c2w_list[idx]
            kf_idx = idx // cfg['back_end']['choose_keyframe_every']
            miou, fwiou, class_avg_accuracy, total_accuracy = slam.novel_view_render(i, gt_color, gt_depth, 
                gt_label, gt_c2w, est_c2w, kf_idx, keyframe_list, keyframe_dict, encoder)

            miou_list.append(miou)
            fwiou_list.append(fwiou)
            class_avg_accuracy_list.append(class_avg_accuracy)
            total_accuracy_list.append(total_accuracy)
            
    
    print(f'Average rendering result: psnr={slam.psnr_sum/count}, ssim={slam.ssim_sum/count}, lpips={slam.lpips_sum/count}')
    print(f'miou={np.mean(miou_list): .4f}, fwiou={np.mean(fwiou_list): .4f}, class_avg_accuracy={np.mean(class_avg_accuracy_list): .4f}, total_accuracy={np.mean(total_accuracy_list): .4f}')
    
    file = open(f'{slam.out_dir}/rendered/rendering_eval.txt', 'a')
    file.write(f'Average: ' +
                f'psnr: {slam.psnr_sum/count:.4f} ' + 
                f'ssim: {slam.ssim_sum/count:.4f} ' +
                f'lpips: {slam.lpips_sum/count:.4f}\n' +
                f'miou: {np.mean(miou_list): .4f}' +
                f'fwiou: {np.mean(fwiou_list): .4f}' +
                f'class_avg_accuracy: {np.mean(class_avg_accuracy_list): .4f}' +
                f'total_accuracy: {np.mean(total_accuracy_list): .4f}\n')
    file.close()
